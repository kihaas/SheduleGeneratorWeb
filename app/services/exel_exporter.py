from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
from datetime import datetime
from typing import List, Dict, Any


class ExcelExporter:
    def __init__(self):
        self.week_days = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
        self.time_slots = [
            '9:00-10:30',
            '10:40-12:10',
            '12:40-14:10',
            '14:20-15:50'
        ]

    async def export_schedule_to_excel(self, lessons: List[Dict[str, Any]], schedule_name: str) -> bytes:
        """Экспорт расписания в Excel"""

        # Создаем workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Расписание"

        # Устанавливаем заголовок
        ws['A1'] = f"Расписание: {schedule_name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Дата генерации
        ws['A2'] = f"Сгенерировано: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
        ws.merge_cells('A2:H2')
        ws['A2'].alignment = Alignment(horizontal='center')

        # Заголовки дней недели
        headers = ['Время'] + self.week_days
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Временные слоты
        for row, time_slot in enumerate(self.time_slots, 5):
            time_cell = ws.cell(row=row, column=1, value=time_slot)
            time_cell.font = Font(bold=True)
            time_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Заполняем расписание
        for lesson in lessons:
            day = lesson.get('day', 0)
            time_slot = lesson.get('time_slot', 0)
            subject = lesson.get('subject_name', '')
            teacher = lesson.get('teacher', '')

            if 0 <= day < 7 and 0 <= time_slot < 4:
                cell_value = f"{subject}\n({teacher})"
                cell = ws.cell(row=time_slot + 5, column=day + 2, value=cell_value)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # Заливка для занятых ячеек
                cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        # Настраиваем ширину колонок
        column_widths = [15, 25, 25, 25, 25, 25, 25, 25]  # A-H
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # Высота строк для временных слотов
        for row in range(5, 9):
            ws.row_dimensions[row].height = 60

        # Сохраняем в bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output.getvalue()


# Глобальный экземпляр
excel_exporter = ExcelExporter()